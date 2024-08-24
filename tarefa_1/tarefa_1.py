import pandas as pd
from openpyxl import load_workbook

# Carregar os dados da planilha
df = pd.read_excel("tarefa_1/vendas.xlsx", sheet_name="Vendas")


# Função para calcular a comissão de cada venda
def calcular_comissao(row):
    comissao = row["Valor da Venda"] * 0.10
    comissao_marketing = 0
    comissao_gerente = 0
    comissao_final = comissao

    # 20% marketing
    if row["Canal de Venda"] == "Online":
        comissao_marketing = comissao * 0.20
        comissao_final -= comissao_marketing

    # Comissão total é maior ou igual a R$ 1.500,00 10% vai para o gerente
    if comissao >= 1500:
        comissao_gerente = comissao * 0.10
        comissao_final -= comissao_gerente

    return comissao, comissao_marketing, comissao_gerente, comissao_final


# Aplicar a função de cálculo de comissão para cada linha
df[["Comissão", "Comissão Marketing", "Comissão Gerente", "Comissão Final"]] = df.apply(
    calcular_comissao, axis=1, result_type="expand"
)

# Agrupar por vendedor para obter o total
resultado = (
    df.groupby("Nome do Vendedor")
    .agg(
        {
            "Comissão": "sum",
            "Comissão Marketing": "sum",
            "Comissão Gerente": "sum",
            "Comissão Final": "sum",
        }
    )
    .reset_index()
)

# Salvar o resultado em uma nova planilha
resultado.to_excel(
    "tarefa_1/resultado_comissoes.xlsx", index=False, sheet_name="Comissão"
)

# Carregar a planilha criada
wb = load_workbook("tarefa_1/resultado_comissoes.xlsx")
ws = wb.active

# Linha de total da comissão de marketing
total_comissao_marketing = df["Comissão Marketing"].sum()
ws.append(["Total", None, total_comissao_marketing, None, None])  # type: ignore

# Formatar as colunas de valores monetários
real_format = "R$ #,##0.00"
for col in ["B", "C", "D", "E"]:
    for cell in ws[col][1:]:  # type: ignore
        cell.number_format = real_format
# Salva a planilha
wb.save("tarefa_1/resultado_comissoes.xlsx")

print("Planilha criada")
