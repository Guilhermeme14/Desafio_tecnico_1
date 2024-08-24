import numpy as np
import pandas as pd
from openpyxl import load_workbook

df_vendas = pd.read_excel("vendas.xlsx", sheet_name="Pagamentos")
df_comissao = pd.read_excel("resultado_comissoes.xlsx", sheet_name="Comissão")

# Mesclar as duas tabelas com base no nome do vendedor
df_juntar = pd.merge(df_comissao, df_vendas, on="Nome do Vendedor", how="outer")

# Tirar a linha Total
df_juntar = df_juntar[df_juntar["Nome do Vendedor"] != "Total"]

# Comparar os valores de comissão calculados com os valores pagos
df_juntar["Valor correto?"] = np.where(
    df_juntar["Comissão Final"] == df_juntar["Comissão_y"], "Correto", "Incorreto"
)

# Selecionar as colunas necessárias para a saída esperada
resultado = df_juntar[
    ["Nome do Vendedor", "Comissão_y", "Valor correto?", "Comissão Final"]
]

resultado = resultado.rename(
    columns={"Comissão_y": "Valor Pago", "Comissão Final": "Valor Correto"}
)

# Salvar o resultado em uma nova planilha
resultado.to_excel("comissao_correta.xlsx", index=False, sheet_name="Comissão correta")

# Carregar a planilha criada
wb = load_workbook("comissao_correta.xlsx")
ws = wb.active

# Formatar as colunas de valores monetários
real_format = "R$ #,##0.00"
for col in ["B", "D"]:
    for cell in ws[col][1:]:  # type: ignore
        cell.number_format = real_format

wb.save("comissao_correta.xlsx")
