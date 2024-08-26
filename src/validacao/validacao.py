import numpy as np
import pandas as pd
from openpyxl import load_workbook


def carregar_planilhas(caminho_vendas, caminho_comissao):
    # Carrega os dados das planilhas de vendas e comissões
    try:
        df_vendas = pd.read_excel(caminho_vendas, sheet_name="Pagamentos")
        df_comissao = pd.read_excel(caminho_comissao, sheet_name="Comissão")
        return df_vendas, df_comissao
    except FileNotFoundError as e:
        print(f"Erro ao carregar arquivos: {e}")
        raise
    except Exception as e:
        print(f"Erro inesperado ao carregar planilhas: {e}")
        raise


def juntar_dados(df_vendas, df_comissao):
    # Mescla as tabelas de vendas e comissões com base no nome do vendedor e compara os valores de comissão
    try:
        df_juntar = pd.merge(df_comissao, df_vendas, on="Nome do Vendedor", how="outer")
        df_juntar = df_juntar[df_juntar["Nome do Vendedor"] != "Total"]
        df_juntar["Valor correto?"] = np.where(
            df_juntar["Comissão Final"] == df_juntar["Comissão_y"],
            "Correto",
            "Incorreto",
        )
        return df_juntar
    except KeyError as e:
        print(f"Erro ao mesclar dados: coluna não encontrada - {e}")
        raise
    except Exception as e:
        print(f"Erro inesperado ao processar dados: {e}")
        raise


def formatar_resultado(df_juntar):
    # Seleciona as colunas necessárias e formata o DataFrame para a saída esperada
    try:
        resultado = df_juntar[
            ["Nome do Vendedor", "Comissão_y", "Valor correto?", "Comissão Final"]
        ]
        resultado = resultado.rename(
            columns={"Comissão_y": "Valor Pago", "Comissão Final": "Valor Correto"}
        )
        return resultado
    except KeyError as e:
        print(f"Erro ao formatar resultado: coluna não encontrada - {e}")
        raise
    except Exception as e:
        print(f"Erro inesperado ao formatar resultado: {e}")
        raise


def salvar_resultado_(resultado, caminho_arquivo):
    # Salva o resultado em uma nova planilha e formata as colunas monetárias
    try:
        resultado.to_excel(caminho_arquivo, index=False, sheet_name="Comissão correta")

        wb = load_workbook(caminho_arquivo)
        ws = wb.active

        real_format = "R$ #,##0.00"
        for col in ["B", "D"]:
            for cell in ws[col][1:]:  # type: ignore
                cell.number_format = real_format

        wb.save(caminho_arquivo)
    except FileNotFoundError as e:
        print(f"Erro ao salvar arquivo: {e}")
        raise
    except Exception as e:
        print(f"Erro inesperado ao salvar planilha: {e}")
        raise
