import pandas as pd
from openpyxl import load_workbook


def carregar_dados(caminho_arquivo, nome_aba):
    # Carrega os dados de uma planilha Excel
    try:
        return pd.read_excel(caminho_arquivo, sheet_name=nome_aba)
    except FileNotFoundError as e:
        print(f"Erro ao carregar arquivo: {e}")
        raise
    except Exception as e:
        print(f"Erro inesperado ao carregar dados: {e}")
        raise


def calcular_comissao(row):
    # Calcula a comissão de cada venda
    try:
        comissao = row["Valor da Venda"] * 0.10
        comissao_marketing = 0
        comissao_gerente = 0
        comissao_final = comissao

        if row["Canal de Venda"] == "Online":
            comissao_marketing = comissao * 0.20
            comissao_final -= comissao_marketing

        if comissao >= 1500:
            comissao_gerente = comissao * 0.10
            comissao_final -= comissao_gerente

        return comissao, comissao_marketing, comissao_gerente, comissao_final
    except KeyError as e:
        print(f"Erro ao calcular comissão: coluna não encontrada - {e}")
        raise
    except Exception as e:
        print(f"Erro inesperado ao calcular comissão: {e}")
        raise


def aplicar_calculo_comissoes(df):
    # Aplica o cálculo de comissão para cada linha do DataFrame e retorna o resultado agrupado por vendedor
    try:
        df[["Comissão", "Comissão Marketing", "Comissão Gerente", "Comissão Final"]] = (
            df.apply(calcular_comissao, axis=1, result_type="expand")
        )

        resultado = (
            df.groupby("Nome do Vendedor")
            .agg(
                {
                    "Comissão": "sum",
                    "Comissão Marketing": "sum",
                    "Comissão Gerente": "sum",
                    "Comissão Final": "sum",
                }
            )
            .reset_index()
        )
        return resultado
    except KeyError as e:
        print(f"Erro ao aplicar cálculo de comissões: coluna não encontrada - {e}")
        raise
    except Exception as e:
        print(f"Erro inesperado ao aplicar cálculo de comissões: {e}")
        raise


def salvar_resultado(df, resultado, caminho_arquivo):
    # Salva o resultado em uma nova planilha e formata as colunas monetárias
    try:
        resultado.to_excel(caminho_arquivo, index=False, sheet_name="Comissão")

        wb = load_workbook(caminho_arquivo)
        ws = wb.active

        total_comissao_marketing = df["Comissão Marketing"].sum()
        ws.append(["Total", None, total_comissao_marketing, None, None])  # type: ignore

        real_format = "R$ #,##0.00"
        for col in ["B", "C", "D", "E"]:
            for cell in ws[col][1:]:  # type: ignore
                cell.number_format = real_format

        wb.save(caminho_arquivo)
    except FileNotFoundError as e:
        print(f"Erro ao salvar arquivo: {e}")
        raise
    except Exception as e:
        print(f"Erro inesperado ao salvar resultado: {e}")
        raise
